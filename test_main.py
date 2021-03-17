# import tensorflow as tf
# import matplotlib.pyplot as plt
import numpy as np
# import models.train_model as tm
import pandas as pd
import os
import scipy.stats as stats
# import config as cfg
# import config_models as cfg_mod
# import data_preprocessing.preprocessing as pp
# import pickle
# from visualization.plotting import plot_forecast
# from data_preprocessing.WindowGenerator import WindowGenerator

# import datetime

# import pvlib forecast models
# from pvlib.forecast import GFS, NAM, NDFD, HRRR, RAP

from openpyxl import load_workbook


def main():
    a = np.random.rand(15, 5)
    sk = stats.skew(a)
    print(sk)
    # a = np.array([range(10)]).T
    # b = np.array([x**2 for x in a])
    # c = np.array([x*5 for x in a])
    # err = np.append(a, b, 1)
    # data_df = pd.DataFrame(err)
    # data_df.columns = ['value', 'square']
    # err1 = np.append(a, c, 1)
    # data_df1 = pd.DataFrame(err1)
    # data_df1.columns = ['value', 'factor5']
    # write_excel(data_df, 'foo')
    # write_excel(data_df1, 'bar')


# def write_excel(df, sheet):
#     # book = load_workbook('saver/outputs/errors/test1.xlsx')
#     # try:
#     #     writer = pd.ExcelWriter('saver/outputs/errors/test1.xlsx', mode='a')
#     # finally:
#     #     writer = pd.ExcelWriter('saver/outputs/errors/test1.xlsx')
#     writepath = 'saver/outputs/errors/test1.xlsx'
#     mode = 'a' if os.path.exists(writepath) else 'w'
#     with pd.ExcelWriter(writepath, mode=mode) as writer:
#         # writer = pd.ExcelWriter('saver/outputs/errors/test1.xlsx', mode='a+')
#         if sheet in writer.book.sheetnames:
#             writer.book.remove(writer.book[sheet])
#         df.to_excel(writer, sheet, float_format='%.5f')
#         writer.save()


# def pvlib_main():
#     models = cfg_mod.models
#     for model in models:
#         df = pd.read_pickle('data/pickles/' + model['city'] + '.pickle')
#         train, val, test, num_features, date_time, column_indices = \
#             pp.preprocess(df, model['fields'], city=model['city'], time=True)
#         model['train'] = train
#         model['val'] = val
#         model['test'] = test
#         model['num_features'] = num_features
#         model['test_datetime'] = date_time
#         model['column_indices'] = column_indices
#
#     latitude, longitude, tz = 32.2, -110.9, 'US/Arizona'
#     # latitude, longitude, tz = 48.39841, 9.99155, 'Europe/Berlin'
#     start = pd.Timestamp(models[0]['test_datetime'][0], tz=tz)
#     end = start + pd.Timedelta(days=365)
#     irrad_vars = ['ghi', 'dni', 'dhi']
#     model = HRRR()
#     data = model.get_processed_data(latitude, longitude, start, end)
#     data[irrad_vars].plot()
#
#     plt.ylabel('Irradiance ($W/m^2$)')
#     plt.xlabel('Forecast Time ({})'.format(tz))
#     plt.title('GFS 3 km forecast for lat={}, lon={}'.format(latitude, longitude))
#
#     plt.legend()
#     plt.show()


if __name__ == '__main__':
    main()
